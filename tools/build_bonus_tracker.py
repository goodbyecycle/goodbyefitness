"""Build the social media subscriber bonus tracker workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLACK = "0A0A0A"      # crushed black (brand)
BONE = "F2EDE4"       # bone white (brand)
PINK = "FF006E"       # graffiti pink (brand accent)
INPUT_FILL = "FFF2CC" # cells the user fills in
BLUE = "0000FF"       # hardcoded inputs
GREEN = "008000"      # links to another sheet

FONT = "Arial"
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

METRICS = [
    # label,                          rate,  unit,                     example prev, example curr
    ("YouTube — Subscribers",         0.50,  "per new subscriber",     1240, 1388),
    ("Instagram — Followers",         0.25,  "per new follower",       3120, 3305),
    ("Facebook — Followers",          0.25,  "per new follower",        870,  902),
    ("Google — Positive Reviews",     1.50,  "per new positive review",  46,   53),
    ("YouTube — Hours Watched",       1.00,  "per additional hour",    2150, 2480),
]
FIRST = 5                       # first data row (same on both sheets)
LAST = FIRST + len(METRICS) - 1
TOTAL_ROW = LAST + 1

wb = Workbook()

# ---------------------------------------------------------------- Sheet 1
ws = wb.active
ws.title = "Monthly Tracker"

ws["A1"] = "Goodbye Coffee — Social Media Bonus Tracker"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=BLACK)
ws.merge_cells("A1:E1")

ws["A2"] = "Previous month:"
ws["B2"] = "July 2026"
ws["C2"] = "Current month:"
ws["D2"] = "August 2026"
for c in ("A2", "C2"):
    ws[c].font = Font(name=FONT, size=10, bold=True)
for c in ("B2", "D2"):
    ws[c].font = Font(name=FONT, size=10, color=BLUE)
    ws[c].fill = PatternFill("solid", fgColor=INPUT_FILL)
    ws[c].border = box

headers = ["Account / Metric", "Previous Month", "Current Month", "Gain", "Bonus"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=i, value=h)
    cell.font = Font(name=FONT, size=11, bold=True, color=BONE)
    cell.fill = PatternFill("solid", fgColor=BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = box
ws.row_dimensions[4].height = 28

for i, (label, _rate, _unit, _p, _c) in enumerate(METRICS):
    r = FIRST + i
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
    for col in (2, 3):                      # user-entered counts
        cell = ws.cell(row=r, column=col)
        cell.font = Font(name=FONT, size=10, color=BLUE)
        cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
        cell.number_format = "#,##0"
    gain = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    gain.font = Font(name=FONT, size=10)
    gain.number_format = "#,##0;(#,##0);-"
    bonus = ws.cell(row=r, column=5, value=f"=MAX(0,D{r})*'Bonus Summary'!$C${r}")
    bonus.font = Font(name=FONT, size=10, color=GREEN)
    bonus.number_format = "$#,##0.00;($#,##0.00);-"
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = box

ws.cell(row=TOTAL_ROW, column=1, value="TOTAL BONUS")
ws.cell(row=TOTAL_ROW, column=5, value=f"=SUM(E{FIRST}:E{LAST})")
ws.cell(row=TOTAL_ROW, column=5).number_format = "$#,##0.00;($#,##0.00);-"
for col in range(1, 6):
    cell = ws.cell(row=TOTAL_ROW, column=col)
    cell.font = Font(name=FONT, size=11, bold=True, color=BONE)
    cell.fill = PatternFill("solid", fgColor=PINK)
    cell.border = box

notes_row = TOTAL_ROW + 2
notes = [
    ("How to use this sheet", True),
    ("1. Type last month's and this month's numbers into the shaded yellow cells (columns B and C).", False),
    ("2. Gain and Bonus calculate themselves — don't type over them.", False),
    ("3. Bonus rates live on the 'Bonus Summary' tab, column C. Change a rate there and both tabs update.", False),
    ("", False),
    ("Bonus rates in force: YouTube subscribers $0.50 · Instagram $0.25 · Facebook $0.25 ·", False),
    ("Google positive reviews $1.50 · YouTube hours watched $1.00 per hour.", False),
    ("", False),
    ("Assumptions (change if wrong):", True),
    ("• Every row is paid on the GAIN over last month, including YouTube hours watched and Google reviews —", False),
    ("  i.e. this month's figure minus last month's, not the month's total.", False),
    ("• A month that goes backwards pays $0, never a negative bonus (MAX(0, gain) in the formula).", False),
    ("• Google reviews counts POSITIVE reviews only; enter the positive-review count, not the overall total.", False),
    ("", False),
    ("Example row (illustration only — not real numbers):", True),
    ("  YouTube — Subscribers · Previous 1,240 · Current 1,388 · Gain 148 · Bonus $74.00", False),
]
for i, (text, bold) in enumerate(notes):
    cell = ws.cell(row=notes_row + i, column=1, value=text)
    cell.font = Font(name=FONT, size=9, bold=bold, italic=not bold and text.startswith("  "))

for col, width in zip("ABCDE", (34, 18, 18, 14, 16)):
    ws.column_dimensions[col].width = width
ws.freeze_panes = "A5"

# ---------------------------------------------------------------- Sheet 2
s2 = wb.create_sheet("Bonus Summary")

s2["A1"] = "Bonus Payout Summary"
s2["A1"].font = Font(name=FONT, size=14, bold=True, color=BLACK)
s2.merge_cells("A1:E1")

s2["A2"] = "Pay period:"
s2["A2"].font = Font(name=FONT, size=10, bold=True)
s2["B2"] = "='Monthly Tracker'!B2&\" → \"&'Monthly Tracker'!D2"
s2["B2"].font = Font(name=FONT, size=10, color=GREEN)

h2 = ["Account / Metric", "Gain", "Rate", "Unit", "Bonus"]
for i, h in enumerate(h2, start=1):
    cell = s2.cell(row=4, column=i, value=h)
    cell.font = Font(name=FONT, size=11, bold=True, color=BONE)
    cell.fill = PatternFill("solid", fgColor=BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = box
s2.row_dimensions[4].height = 28

for i, (label, rate, unit, _p, _c) in enumerate(METRICS):
    r = FIRST + i
    a = s2.cell(row=r, column=1, value=f"='Monthly Tracker'!A{r}")
    a.font = Font(name=FONT, size=10, color=GREEN)
    g = s2.cell(row=r, column=2, value=f"='Monthly Tracker'!D{r}")
    g.font = Font(name=FONT, size=10, color=GREEN)
    g.number_format = "#,##0;(#,##0);-"
    rt = s2.cell(row=r, column=3, value=rate)          # the one place rates are set
    rt.font = Font(name=FONT, size=10, color=BLUE)
    rt.fill = PatternFill("solid", fgColor=INPUT_FILL)
    rt.number_format = "$0.00"
    u = s2.cell(row=r, column=4, value=unit)
    u.font = Font(name=FONT, size=10)
    b = s2.cell(row=r, column=5, value=f"='Monthly Tracker'!E{r}")
    b.font = Font(name=FONT, size=10, color=GREEN)
    b.number_format = "$#,##0.00;($#,##0.00);-"
    for col in range(1, 6):
        s2.cell(row=r, column=col).border = box

s2.cell(row=TOTAL_ROW, column=1, value="TOTAL BONUS OWED")
s2.cell(row=TOTAL_ROW, column=5, value=f"=SUM(E{FIRST}:E{LAST})")
s2.cell(row=TOTAL_ROW, column=5).number_format = "$#,##0.00;($#,##0.00);-"
for col in range(1, 6):
    cell = s2.cell(row=TOTAL_ROW, column=col)
    cell.font = Font(name=FONT, size=11, bold=True, color=BONE)
    cell.fill = PatternFill("solid", fgColor=PINK)
    cell.border = box

p = TOTAL_ROW + 2
payout = [
    ("Payout detail", ""),
    ("Total gain across follower/subscriber rows", f"=SUM(B{FIRST}:B{FIRST+2})"),
    ("Bonus from followers & subscribers", f"=SUM(E{FIRST}:E{FIRST+2})"),
    ("Bonus from YouTube hours watched", f"=E{LAST}"),
    ("Bonus from Google positive reviews", f"=E{LAST-1}"),
    ("Total bonus owed this period", f"=E{TOTAL_ROW}"),
]
for i, (label, formula) in enumerate(payout):
    r = p + i
    lc = s2.cell(row=r, column=1, value=label)
    lc.font = Font(name=FONT, size=10, bold=(i == 0 or i == len(payout) - 1))
    if formula:
        vc = s2.cell(row=r, column=2, value=formula)
        vc.font = Font(name=FONT, size=10, bold=(i == len(payout) - 1))
        vc.number_format = "#,##0;(#,##0);-" if i == 1 else "$#,##0.00;($#,##0.00);-"
        vc.border = box

n = p + len(payout) + 2
notes2 = [
    ("Notes", True),
    ("• Column C is the only place bonus rates are set — the Monthly Tracker tab reads them from here.", False),
    ("• Everything else on this tab is pulled from the Monthly Tracker; enter data there, not here.", False),
    ("• Bonus rates in force: YouTube subs $0.50, Instagram $0.25, Facebook $0.25,", False),
    ("  Google positive reviews $1.50, YouTube hours watched $1.00 per hour.", False),
    ("• Bonus is paid on the month-over-month gain and is floored at $0 for any row that declines.", False),
]
for i, (text, bold) in enumerate(notes2):
    s2.cell(row=n + i, column=1, value=text).font = Font(name=FONT, size=9, bold=bold)

for col, width in zip("ABCDE", (34, 14, 12, 26, 16)):
    s2.column_dimensions[col].width = width
s2.freeze_panes = "A5"

out = "/home/user/goodbyefitness/social-media-bonus-tracker.xlsx"
wb.save(out)
print("wrote", out)
